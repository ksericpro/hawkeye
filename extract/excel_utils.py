
from openpyxl import load_workbook
from dateutil import parser

def is_valid_timestamp_parse(timestamp_str):
    try:
        parser.parse(timestamp_str)
        return True
    except ValueError:
        return False

# date_str = "06/01/2024 2024"
# output_format = "%d/%m/%Y"
# dt_object = parser.parse(date_str,dayfirst=True)
# print(dt_object.strftime(output_format))

def split_period(time_period, date_format="%d/%m/%Y"):
    print('time period: ', time_period)

    if is_valid_timestamp_parse(time_period):
        start_date = parser.parse(time_period, dayfirst=True).strftime(date_format)
        return start_date, start_date
    else:
        strs = time_period.split("-")
        if len(strs) == 2:
            start_d = strs[0]
            end_d = strs[1]
            if is_valid_timestamp_parse(start_d) and is_valid_timestamp_parse(end_d):
                start_date = parser.parse(start_d, dayfirst=True).strftime(date_format)
                end_date = parser.parse(end_d, dayfirst=True).strftime(date_format)
                return start_date, end_date
        else:
            return None, None

# print(split_period("2024-02-25 00:00:00"))

workbook = load_workbook(filename="CDR_excel/cdr_excel.xlsx")
# sheet = workbook.active
sheet = workbook['Sheet 1']

# define cell for each sub tables
cell_loc=[(2,4,1,2), (2,4,4,5),(19,21,1,2),(8,17,1,5)]

def extract_info(workbook, cell_loc):
    result_json={}
    for item in cell_loc:
        if (item[1] - item[0]) < 5:
            for row in sheet.iter_rows(min_row=item[0],
                                       max_row=item[1],
                                       min_col=item[2],
                                       max_col=item[3],
                                       values_only=True):

                result_json[row[0]] = row[1]
        else:
            requests = {}
            for row in sheet.iter_rows(min_row=item[0],
                                       max_row=item[1],
                                       min_col=item[2],
                                       max_col=item[3],
                                       values_only=True):
                s_no = row[0]
                period = (row[4])
                start_date, end_date = split_period(str(period))
                print('------', start_date, end_date)

                if row[1] or row[2]:
                    request_ = {
                        # "S/N": row[0],
                        "toi": row[1],
                        "telco": row[2],
                        "type": row[3],
                        "Request Period": row[4],
                        "cdrRequestPeriodStart": start_date,
                        "cdrRequestPeriodEnd": end_date
                    }
                    requests[s_no] = request_

            result_json['Requests'] = requests
    result_json['pidJobNo'] = result_json.pop('Xxx Job No:')
    result_json['caseDescription'] = result_json.pop('Case Description')
    result_json['projectName'] = result_json.pop('Project Name')
    result_json['dataToWhom'] = result_json.pop('Requesting Officer')
    return result_json

result_json = extract_info(workbook, cell_loc)

print(result_json)
